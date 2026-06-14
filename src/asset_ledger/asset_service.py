from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, replace
from datetime import date, datetime
from typing import Any
from uuid import uuid4

from openpyxl.utils import get_column_letter

from .excel_repository import (
    ASSET_HEADERS,
    ASSET_SHEET,
    CATEGORY_SHEET,
    CHANGE_SHEET,
    CONFIG_SHEET,
    ENUM_SHEET,
    LOCATION_SHEET,
    STORAGE_HEADERS,
    STORAGE_SHEET,
    ExcelRepository,
    WorkbookChangedExternallyError,
)
from .models import Asset, ChangeRecord, DictionaryItem, StorageMedium, timestamp_now
from .snapshot import CacheStatus, LedgerSnapshot


class AssetServiceError(ValueError):
    pass


class DuplicateEquipmentCodeError(AssetServiceError):
    pass


class AssetNotFoundError(AssetServiceError):
    pass


class CacheUnavailableError(AssetServiceError):
    pass


class CacheReloadAfterSaveError(AssetServiceError):
    pass


FIELD_TO_HEADER = {
    "asset_id": "资产唯一标识符",
    "equipment_code": "bm编码",
    "name": "设备器材",
    "primary_category": "一级类别",
    "secondary_category": "二级类别",
    "product_spec": "产品规格",
    "model": "产品型号",
    "serial_number": "设备序列号",
    "manufacturer": "生产厂家",
    "supplier": "供应商",
    "brand": "品牌",
    "source": "取得方式",
    "purchase_date": "取得日期",
    "start_date": "启用日期",
    "valuation_method": "计价方式",
    "price": "金额",
    "department": "管理部门",
    "owner": "管理人",
    "use_department": "使用部门",
    "user": "使用人",
    "location": "存放地点",
    "manufacture_date": "出厂日期",
    "grade": "等级",
    "status": "使用状态",
    "label_printed": "是否已打印标签",
    "notes": "备注",
    "created_at": "创建时间",
    "updated_at": "更新时间",
}
HEADER_TO_FIELD = {header: field for field, header in FIELD_TO_HEADER.items()}
EDITABLE_FIELDS = [
    field
    for field in FIELD_TO_HEADER
    if field not in {"asset_id", "created_at", "updated_at"}
]

STORAGE_FIELD_TO_HEADER = {
    "storage_id": "存储介质ID",
    "asset_id": "资产唯一标识符",
    "medium_type": "介质类型",
    "name": "名称/编号",
    "brand": "品牌",
    "capacity_value": "容量数值",
    "capacity_unit": "容量单位",
    "model": "型号",
    "serial_number": "序列号",
    "notes": "备注",
    "created_at": "创建时间",
    "updated_at": "更新时间",
}
STORAGE_HEADER_TO_FIELD = {
    header: field for field, header in STORAGE_FIELD_TO_HEADER.items()
}
STORAGE_EDITABLE_FIELDS = [
    field
    for field in STORAGE_FIELD_TO_HEADER
    if field not in {"storage_id", "asset_id", "created_at", "updated_at"}
]


class AssetService:
    def __init__(self, repository: ExcelRepository) -> None:
        self.repository = repository
        self._snapshot: LedgerSnapshot | None = None
        self.cache_status = CacheStatus()
        self.reload_cache()

    def reload_cache(self) -> CacheStatus:
        last_error: Exception | None = None
        for _attempt in range(2):
            workbook = None
            try:
                workbook, fingerprint = self.repository.load_with_fingerprint(data_only=True)
                snapshot = self._snapshot_from_workbook(workbook, fingerprint)
                self._snapshot = snapshot
                self.cache_status = CacheStatus(loaded_at=snapshot.loaded_at, stale=False)
                return self.cache_status
            except WorkbookChangedExternallyError as exc:
                last_error = exc
            except Exception as exc:
                self._mark_cache_stale(exc)
                raise
            finally:
                if workbook is not None:
                    workbook.close()
        self._mark_cache_stale(last_error)
        raise last_error or CacheUnavailableError("无法加载 Excel 缓存")

    def list_assets(
        self, filters: dict[str, str | None] | None = None, search_text: str = ""
    ) -> list[Asset]:
        snapshot = self._require_snapshot()
        filters = filters or {}
        filtered: list[Asset] = []
        needle = search_text.strip().lower()
        searchable_fields = (
            "asset_id",
            "equipment_code",
            "name",
            "product_spec",
            "brand",
            "model",
            "serial_number",
            "manufacturer",
            "supplier",
            "department",
            "owner",
            "use_department",
            "user",
        )
        for asset in snapshot.assets:
            if needle and not any(
                needle in str(getattr(asset, field) or "").lower()
                for field in searchable_fields
            ) and needle not in snapshot.storage_search_text.get(asset.asset_id, ""):
                continue
            if not self._matches_filters(asset, filters):
                continue
            filtered.append(replace(asset))
        return sorted(filtered, key=lambda item: item.updated_at, reverse=True)

    def get_asset(self, asset_id: str) -> Asset:
        asset = self._require_snapshot().assets_by_id.get(asset_id)
        if not asset:
            raise AssetNotFoundError(f"未找到设备：{asset_id}")
        return replace(asset)

    def create_asset(
        self,
        asset_data: dict[str, Any] | Asset,
        storage_media: list[dict[str, Any] | StorageMedium] | str | None = None,
        change_note: str = "",
    ) -> Asset:
        if isinstance(storage_media, str) and not change_note:
            change_note = storage_media
            storage_media = None
        asset = self._normalize_asset(asset_data)
        media = self._normalize_storage_media(storage_media or [])
        self._validate_asset(asset)
        self._validate_storage_media(media)
        workbook = self._load_for_write()
        try:
            asset_sheet = workbook[ASSET_SHEET]
            self._ensure_equipment_code_unique(asset_sheet, asset.equipment_code)
            asset.asset_id = self._next_asset_id(workbook)
            now = timestamp_now()
            asset.created_at = now
            asset.updated_at = now
            event_group_id = self._new_id("EVT")
            self._append_or_replace_blank(asset_sheet, self._asset_to_row(asset))
            self._format_asset_row(asset_sheet, asset_sheet.max_row)
            self._resize_table(asset_sheet)
            self._append_change(
                workbook,
                ChangeRecord(
                    change_id=self._new_id("CHG"),
                    event_group_id=event_group_id,
                    asset_id=asset.asset_id,
                    event_type="新建设备",
                    field_name="*",
                    old_value="",
                    new_value=asset.name,
                    changed_at=now,
                    operator=self._get_config_value(workbook, "默认操作者", "管理员"),
                    note=change_note,
                ),
            )
            storage_sheet = workbook[STORAGE_SHEET]
            for medium in media:
                medium.storage_id = self._new_id("MED")
                medium.asset_id = asset.asset_id
                medium.created_at = now
                medium.updated_at = now
                self._append_or_replace_blank(storage_sheet, self._storage_to_row(medium))
                self._format_storage_row(storage_sheet, storage_sheet.max_row)
                self._append_storage_change(
                    workbook,
                    medium,
                    "存储介质新增",
                    "存储介质",
                    "",
                    self._storage_summary(medium),
                    event_group_id,
                    now,
                    change_note,
                )
            self._resize_table(storage_sheet)
            self._save_and_reload(workbook)
        finally:
            workbook.close()
        return self.get_asset(asset.asset_id)

    def update_asset(
        self,
        asset_id: str,
        asset_data: dict[str, Any] | Asset,
        storage_media: list[dict[str, Any] | StorageMedium] | str | None = None,
        change_note: str = "",
    ) -> Asset:
        if isinstance(storage_media, str) and not change_note:
            change_note = storage_media
            storage_media = None
        candidate = self._normalize_asset(asset_data)
        candidate_media = (
            self._normalize_storage_media(storage_media)
            if storage_media is not None
            else None
        )
        self._validate_asset(candidate)
        if candidate_media is not None:
            self._validate_storage_media(candidate_media)
        workbook = self._load_for_write()
        try:
            sheet = workbook[ASSET_SHEET]
            row_number = self._find_asset_row(sheet, asset_id)
            if row_number is None:
                raise AssetNotFoundError(f"未找到设备：{asset_id}")
            current = self._asset_from_row(
                tuple(
                    sheet.cell(row=row_number, column=index).value
                    for index in range(1, len(ASSET_HEADERS) + 1)
                )
            )
            self._ensure_equipment_code_unique(
                sheet, candidate.equipment_code, excluded_asset_id=asset_id
            )
            candidate.asset_id = current.asset_id
            candidate.created_at = current.created_at
            candidate.updated_at = current.updated_at
            changes = [
                field
                for field in EDITABLE_FIELDS
                if self._comparable(getattr(current, field))
                != self._comparable(getattr(candidate, field))
            ]
            current_media = [
                self._storage_from_row(row)
                for row in workbook[STORAGE_SHEET].iter_rows(min_row=2, values_only=True)
                if row[0] and str(row[1]) == asset_id
            ]
            storage_changes = self._storage_changes(
                current_media, candidate_media if candidate_media is not None else current_media
            )
            if not changes and not storage_changes:
                return current

            now = timestamp_now()
            candidate.updated_at = now
            for column, value in enumerate(self._asset_to_row(candidate), start=1):
                sheet.cell(row=row_number, column=column).value = value
            self._format_asset_row(sheet, row_number)

            event_group_id = self._new_id("EVT")
            operator = self._get_config_value(workbook, "默认操作者", "管理员")
            for field in changes:
                header = FIELD_TO_HEADER[field]
                event_type = (
                    "使用状态变更"
                    if field == "status"
                    else "存放地点变更"
                    if field == "location"
                    else "信息修改"
                )
                self._append_change(
                    workbook,
                    ChangeRecord(
                        change_id=self._new_id("CHG"),
                        event_group_id=event_group_id,
                        asset_id=asset_id,
                        event_type=event_type,
                        field_name=header,
                        old_value=self._display_value(getattr(current, field)),
                        new_value=self._display_value(getattr(candidate, field)),
                        changed_at=now,
                        operator=operator,
                        note=change_note,
                    ),
                )
            if candidate_media is not None:
                self._replace_storage_media(
                    workbook,
                    asset_id,
                    candidate_media,
                    storage_changes,
                    event_group_id,
                    now,
                    operator,
                    change_note,
                )
            self._save_and_reload(workbook)
        finally:
            workbook.close()
        return self.get_asset(asset_id)

    def list_changes(self, asset_id: str, field_name: str = "") -> list[ChangeRecord]:
        records = self._require_snapshot().changes_by_asset_id.get(asset_id, ())
        return [
            replace(change)
            for change in records
            if not field_name or change.field_name == field_name
        ]

    def list_storage_media(self, asset_id: str) -> list[StorageMedium]:
        media = self._require_snapshot().storage_by_asset_id.get(asset_id, ())
        return [replace(item) for item in media]

    def list_departments(self) -> list[str]:
        return sorted({asset.department for asset in self.list_assets() if asset.department})

    def list_owners(self) -> list[str]:
        return sorted({asset.owner for asset in self.list_assets() if asset.owner})

    def list_use_departments(self) -> list[str]:
        return sorted(
            {asset.use_department for asset in self.list_assets() if asset.use_department}
        )

    def list_users(self) -> list[str]:
        return sorted({asset.user for asset in self.list_assets() if asset.user})

    def list_grades(self) -> list[str]:
        return sorted({asset.grade for asset in self.list_assets() if asset.grade})

    def list_valuation_methods(self) -> list[str]:
        return sorted(
            {asset.valuation_method for asset in self.list_assets() if asset.valuation_method}
        )

    def get_categories(self, *, include_disabled: bool = False) -> list[DictionaryItem]:
        return self._get_dictionary_items(CATEGORY_SHEET, include_disabled)

    def get_locations(self, *, include_disabled: bool = False) -> list[DictionaryItem]:
        return self._get_dictionary_items(LOCATION_SHEET, include_disabled)

    def get_enum_values(self, enum_type: str, *, include_disabled: bool = False) -> list[str]:
        values = self._require_snapshot().enum_values.get(enum_type, ())
        return [
            value
            for _order, value, enabled in values
            if include_disabled or enabled
        ]

    def get_operator(self) -> str:
        return self._require_snapshot().config.get("默认操作者", "管理员")

    def set_operator(self, operator: str) -> None:
        workbook = self._load_for_write()
        try:
            self._set_config_value(workbook, "默认操作者", operator.strip() or "管理员")
            self._save_and_reload(workbook)
        finally:
            workbook.close()

    def add_dictionary_item(
        self, sheet_name: str, name: str, *, parent_id: str = ""
    ) -> DictionaryItem:
        if sheet_name not in {CATEGORY_SHEET, LOCATION_SHEET}:
            raise AssetServiceError(f"不支持的字典：{sheet_name}")
        clean_name = name.strip()
        if not clean_name:
            raise AssetServiceError("名称不能为空")
        workbook = self._load_for_write()
        try:
            sheet = workbook[sheet_name]
            existing_rows = [
                row
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0]
            ]
            if any(
                str(row[2]) == clean_name and str(row[1] or "") == parent_id
                for row in existing_rows
            ):
                raise AssetServiceError(f"字典项已存在：{clean_name}")
            order = max((int(row[3] or 0) for row in existing_rows), default=0) + 10
            prefix = "CAT" if sheet_name == CATEGORY_SHEET else "LOC"
            item = DictionaryItem(
                item_id=self._new_id(prefix),
                parent_id=parent_id,
                name=clean_name,
                order=order,
                enabled=True,
            )
            self._append_or_replace_blank(
                sheet, [item.item_id, item.parent_id, item.name, item.order, item.enabled]
            )
            self._resize_table(sheet)
            self._save_and_reload(workbook)
        finally:
            workbook.close()
        return item

    def set_dictionary_item_enabled(
        self, sheet_name: str, item_id: str, enabled: bool
    ) -> None:
        if sheet_name not in {CATEGORY_SHEET, LOCATION_SHEET}:
            raise AssetServiceError(f"不支持的字典：{sheet_name}")
        workbook = self._load_for_write()
        try:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == item_id:
                    row[4].value = bool(enabled)
                    self._save_and_reload(workbook)
                    return
            raise AssetServiceError(f"未找到字典项：{item_id}")
        finally:
            workbook.close()

    def set_enum_values(self, enum_type: str, values: list[str]) -> None:
        cleaned = list(dict.fromkeys(value.strip() for value in values if value.strip()))
        if not cleaned:
            raise AssetServiceError(f"{enum_type}至少需要一个选项")
        workbook = self._load_for_write()
        try:
            sheet = workbook[ENUM_SHEET]
            preserved = [
                list(row)
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] and row[0] != enum_type
            ]
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)
            for row in preserved:
                sheet.append(row)
            for index, value in enumerate(cleaned, start=1):
                sheet.append([enum_type, value, index * 10, True])
            self._resize_table(sheet)
            self._save_and_reload(workbook)
        finally:
            workbook.close()

    def _get_dictionary_items(
        self, sheet_name: str, include_disabled: bool
    ) -> list[DictionaryItem]:
        items = self._require_snapshot().dictionaries.get(sheet_name, ())
        return [
            replace(item)
            for item in items
            if include_disabled or item.enabled
        ]

    def _snapshot_from_workbook(self, workbook, fingerprint) -> LedgerSnapshot:
        assets = tuple(
            self._asset_from_row(row)
            for row in workbook[ASSET_SHEET].iter_rows(min_row=2, values_only=True)
            if row[0]
        )
        storage_by_asset_id: dict[str, list[StorageMedium]] = defaultdict(list)
        storage_search_text: dict[str, str] = {}
        for row in workbook[STORAGE_SHEET].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            medium = self._storage_from_row(row)
            storage_by_asset_id[medium.asset_id].append(medium)
            storage_search_text[medium.asset_id] = " ".join(
                [storage_search_text.get(medium.asset_id, "")]
                + [str(value or "") for value in row[2:10]]
            ).lower()

        changes_by_asset_id: dict[str, list[ChangeRecord]] = defaultdict(list)
        for row in workbook[CHANGE_SHEET].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            change = ChangeRecord(
                change_id=str(row[0] or ""),
                event_group_id=str(row[1] or ""),
                asset_id=str(row[2] or ""),
                event_type=str(row[3] or ""),
                field_name=str(row[4] or ""),
                old_value=self._display_value(row[5]),
                new_value=self._display_value(row[6]),
                changed_at=self._display_value(row[7]),
                operator=str(row[8] or ""),
                note=str(row[9] or ""),
            )
            changes_by_asset_id[change.asset_id].append(change)

        dictionaries: dict[str, tuple[DictionaryItem, ...]] = {}
        for sheet_name in (CATEGORY_SHEET, LOCATION_SHEET):
            items = [
                DictionaryItem(
                    item_id=str(row[0] or ""),
                    parent_id=str(row[1] or ""),
                    name=str(row[2] or ""),
                    order=int(row[3] or 0),
                    enabled=bool(row[4]),
                )
                for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True)
                if row[0]
            ]
            dictionaries[sheet_name] = tuple(sorted(items, key=lambda item: item.order))

        enum_values: dict[str, list[tuple[int, str, bool]]] = defaultdict(list)
        for row in workbook[ENUM_SHEET].iter_rows(min_row=2, values_only=True):
            if row[0]:
                enum_values[str(row[0])].append(
                    (int(row[2] or 0), str(row[1] or ""), bool(row[3]))
                )
        config = {
            str(row[0]): str(row[1] or "")
            for row in workbook[CONFIG_SHEET].iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        loaded_at = timestamp_now()
        return LedgerSnapshot(
            assets=assets,
            assets_by_id={asset.asset_id: asset for asset in assets},
            storage_by_asset_id={
                asset_id: tuple(
                    sorted(items, key=lambda item: (item.created_at, item.storage_id))
                )
                for asset_id, items in storage_by_asset_id.items()
            },
            changes_by_asset_id={
                asset_id: tuple(
                    sorted(items, key=lambda item: item.changed_at, reverse=True)
                )
                for asset_id, items in changes_by_asset_id.items()
            },
            storage_search_text=storage_search_text,
            dictionaries=dictionaries,
            enum_values={
                enum_type: tuple(sorted(items))
                for enum_type, items in enum_values.items()
            },
            config=config,
            fingerprint=fingerprint,
            loaded_at=loaded_at,
        )

    def _require_snapshot(self) -> LedgerSnapshot:
        if self._snapshot is None:
            raise CacheUnavailableError("内存缓存尚未加载，请点击刷新后重试。")
        return self._snapshot

    def _require_writable_snapshot(self) -> LedgerSnapshot:
        snapshot = self._require_snapshot()
        if self.cache_status.stale:
            raise CacheUnavailableError(
                "内存缓存已过期，为避免覆盖数据，已禁止保存。请点击刷新后重试。"
            )
        return snapshot

    def _load_for_write(self):
        snapshot = self._require_writable_snapshot()
        try:
            return self.repository.load(expected_fingerprint=snapshot.fingerprint)
        except WorkbookChangedExternallyError as exc:
            self._mark_cache_stale(exc)
            raise

    def _save_and_reload(self, workbook) -> None:
        snapshot = self._require_writable_snapshot()
        try:
            self.repository.save(
                workbook,
                expected_fingerprint=snapshot.fingerprint,
            )
        except WorkbookChangedExternallyError as exc:
            self._mark_cache_stale(exc)
            raise
        try:
            self.reload_cache()
        except Exception as exc:
            self._mark_cache_stale(exc)
            raise CacheReloadAfterSaveError(
                "Excel 已保存，但内存缓存刷新失败。为避免重复写入，已禁止继续保存，"
                "请点击刷新或重新启动程序。"
            ) from exc

    def _mark_cache_stale(self, error: Exception | None) -> None:
        loaded_at = self._snapshot.loaded_at if self._snapshot else ""
        self.cache_status = CacheStatus(
            loaded_at=loaded_at,
            stale=True,
            error=str(error or "缓存状态未知"),
        )

    @staticmethod
    def _matches_filters(asset: Asset, filters: dict[str, str | None]) -> bool:
        for key, value in filters.items():
            if value == "":
                continue
            asset_value = str(getattr(asset, key, "") or "")
            if value is None:
                if asset_value:
                    return False
            elif asset_value != str(value):
                return False
        return True

    @staticmethod
    def _normalize_asset(asset_data: dict[str, Any] | Asset) -> Asset:
        if isinstance(asset_data, Asset):
            return Asset.from_mapping(asdict(asset_data))
        return Asset.from_mapping(asset_data)

    @staticmethod
    def _normalize_storage_media(
        storage_media: list[dict[str, Any] | StorageMedium],
    ) -> list[StorageMedium]:
        normalized = []
        for item in storage_media:
            medium = (
                StorageMedium.from_mapping(asdict(item))
                if isinstance(item, StorageMedium)
                else StorageMedium.from_mapping(item)
            )
            if any(
                AssetService._comparable(getattr(medium, field))
                for field in STORAGE_EDITABLE_FIELDS
            ):
                normalized.append(medium)
        return normalized

    @staticmethod
    def _validate_asset(asset: Asset) -> None:
        if not asset.name.strip():
            raise AssetServiceError("设备器材不能为空")
        if asset.price < 0:
            raise AssetServiceError("金额不能为负数")

    @staticmethod
    def _validate_storage_media(storage_media: list[StorageMedium]) -> None:
        storage_ids = [item.storage_id for item in storage_media if item.storage_id]
        if len(storage_ids) != len(set(storage_ids)):
            raise AssetServiceError("存储介质ID不能重复")
        if any(item.capacity_value < 0 for item in storage_media):
            raise AssetServiceError("存储介质容量不能为负数")

    @staticmethod
    def _ensure_equipment_code_unique(
        sheet, equipment_code: str, excluded_asset_id: str = ""
    ) -> None:
        if not equipment_code.strip():
            return
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]) != excluded_asset_id and str(row[1]) == equipment_code:
                raise DuplicateEquipmentCodeError(f"bm编码已存在：{equipment_code}")

    @staticmethod
    def _find_asset_row(sheet, asset_id: str) -> int | None:
        for row_number in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_number, column=1).value == asset_id:
                return row_number
        return None

    @staticmethod
    def _next_asset_id(workbook) -> str:
        value = int(AssetService._get_config_value(workbook, "下一个资产序号", "1"))
        existing_ids = {
            str(row[0])
            for row in workbook[ASSET_SHEET].iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        candidate = f"AST-{datetime.now().year}-{value:06d}"
        while candidate in existing_ids:
            value += 1
            candidate = f"AST-{datetime.now().year}-{value:06d}"
        AssetService._set_config_value(workbook, "下一个资产序号", str(value + 1))
        return candidate

    @staticmethod
    def _get_config_value(workbook, key: str, default: str = "") -> str:
        for row in workbook[CONFIG_SHEET].iter_rows(min_row=2):
            if row[0].value == key:
                return str(row[1].value or default)
        return default

    @staticmethod
    def _set_config_value(workbook, key: str, value: str) -> None:
        sheet = workbook[CONFIG_SHEET]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == key:
                row[1].value = value
                return
        sheet.append([key, value, ""])
        AssetService._resize_table(sheet)

    @staticmethod
    def _append_change(workbook, change: ChangeRecord) -> None:
        sheet = workbook[CHANGE_SHEET]
        AssetService._append_or_replace_blank(
            sheet,
            [
                change.change_id,
                change.event_group_id,
                change.asset_id,
                change.event_type,
                change.field_name,
                change.old_value,
                change.new_value,
                change.changed_at,
                change.operator,
                change.note,
            ],
        )
        AssetService._format_change_row(sheet, sheet.max_row)
        AssetService._resize_table(sheet)

    @staticmethod
    def _append_storage_change(
        workbook,
        medium: StorageMedium,
        event_type: str,
        field_name: str,
        old_value: str,
        new_value: str,
        event_group_id: str,
        changed_at: str,
        note: str,
    ) -> None:
        AssetService._append_change(
            workbook,
            ChangeRecord(
                change_id=AssetService._new_id("CHG"),
                event_group_id=event_group_id,
                asset_id=medium.asset_id,
                event_type=event_type,
                field_name=field_name,
                old_value=old_value,
                new_value=new_value,
                changed_at=changed_at,
                operator=AssetService._get_config_value(workbook, "默认操作者", "管理员"),
                note=note,
            ),
        )

    @staticmethod
    def _storage_changes(
        current_media: list[StorageMedium], candidate_media: list[StorageMedium]
    ) -> list[tuple[str, StorageMedium | None, StorageMedium | None, list[str]]]:
        current_by_id = {item.storage_id: item for item in current_media}
        candidate_ids = {item.storage_id for item in candidate_media if item.storage_id}
        changes: list[tuple[str, StorageMedium | None, StorageMedium | None, list[str]]] = []
        for candidate in candidate_media:
            current = current_by_id.get(candidate.storage_id)
            if not current:
                changes.append(("存储介质新增", None, candidate, []))
                continue
            changed_fields = [
                field
                for field in STORAGE_EDITABLE_FIELDS
                if AssetService._comparable(getattr(current, field))
                != AssetService._comparable(getattr(candidate, field))
            ]
            if changed_fields:
                changes.append(("存储介质修改", current, candidate, changed_fields))
        for current in current_media:
            if current.storage_id not in candidate_ids:
                changes.append(("存储介质移除", current, None, []))
        return changes

    @staticmethod
    def _replace_storage_media(
        workbook,
        asset_id: str,
        candidate_media: list[StorageMedium],
        changes: list[tuple[str, StorageMedium | None, StorageMedium | None, list[str]]],
        event_group_id: str,
        now: str,
        operator: str,
        note: str,
    ) -> None:
        del operator
        sheet = workbook[STORAGE_SHEET]
        current_by_id = {
            item.storage_id: item
            for item in (
                AssetService._storage_from_row(row)
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] and str(row[1]) == asset_id
            )
        }
        for row_number in range(sheet.max_row, 1, -1):
            if str(sheet.cell(row=row_number, column=2).value or "") == asset_id:
                sheet.delete_rows(row_number)

        for medium in candidate_media:
            current = current_by_id.get(medium.storage_id)
            if not medium.storage_id:
                medium.storage_id = AssetService._new_id("MED")
            medium.asset_id = asset_id
            medium.created_at = current.created_at if current else now
            medium.updated_at = (
                now
                if not current
                or any(
                    AssetService._comparable(getattr(current, field))
                    != AssetService._comparable(getattr(medium, field))
                    for field in STORAGE_EDITABLE_FIELDS
                )
                else current.updated_at
            )
            AssetService._append_or_replace_blank(sheet, AssetService._storage_to_row(medium))
            AssetService._format_storage_row(sheet, sheet.max_row)
        if sheet.max_row == 1:
            sheet.append([None] * len(STORAGE_HEADERS))
        AssetService._resize_table(sheet)

        for event_type, current, candidate, fields in changes:
            medium = candidate or current
            if not medium:
                continue
            medium.asset_id = asset_id
            if event_type == "存储介质新增":
                AssetService._append_storage_change(
                    workbook,
                    medium,
                    event_type,
                    "存储介质",
                    "",
                    AssetService._storage_summary(medium),
                    event_group_id,
                    now,
                    note,
                )
            elif event_type == "存储介质移除":
                AssetService._append_storage_change(
                    workbook,
                    medium,
                    event_type,
                    "存储介质",
                    AssetService._storage_summary(medium),
                    "",
                    event_group_id,
                    now,
                    note,
                )
            else:
                for field in fields:
                    AssetService._append_storage_change(
                        workbook,
                        medium,
                        event_type,
                        f"存储介质-{STORAGE_FIELD_TO_HEADER[field]}",
                        AssetService._display_value(getattr(current, field)),
                        AssetService._display_value(getattr(candidate, field)),
                        event_group_id,
                        now,
                        note,
                    )

    @staticmethod
    def _append_or_replace_blank(sheet, values: list[Any]) -> None:
        if sheet.max_row == 2 and not any(
            sheet.cell(row=2, column=column).value for column in range(1, sheet.max_column + 1)
        ):
            for column, value in enumerate(values, start=1):
                sheet.cell(row=2, column=column).value = value
        else:
            sheet.append(values)

    @staticmethod
    def _resize_table(sheet) -> None:
        if not sheet.tables:
            return
        table = next(iter(sheet.tables.values()))
        table.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        sheet.auto_filter.ref = table.ref

    @staticmethod
    def _asset_to_row(asset: Asset) -> list[Any]:
        row: list[Any] = []
        for header in ASSET_HEADERS:
            field = HEADER_TO_FIELD[header]
            value = getattr(asset, field)
            if field in {"purchase_date", "start_date", "manufacture_date"} and value:
                try:
                    value = date.fromisoformat(str(value))
                except ValueError:
                    pass
            row.append(value)
        return row

    @staticmethod
    def _asset_from_row(row: tuple[Any, ...]) -> Asset:
        values = {
            HEADER_TO_FIELD[header]: row[index] if index < len(row) else ""
            for index, header in enumerate(ASSET_HEADERS)
        }
        for key in values:
            if values[key] is None:
                values[key] = ""
        return Asset.from_mapping(values)

    @staticmethod
    def _storage_to_row(medium: StorageMedium) -> list[Any]:
        return [
            getattr(medium, STORAGE_HEADER_TO_FIELD[header])
            for header in STORAGE_HEADERS
        ]

    @staticmethod
    def _storage_from_row(row: tuple[Any, ...]) -> StorageMedium:
        values = {
            STORAGE_HEADER_TO_FIELD[header]: row[index] if index < len(row) else ""
            for index, header in enumerate(STORAGE_HEADERS)
        }
        return StorageMedium.from_mapping(
            {key: "" if value is None else value for key, value in values.items()}
        )

    @staticmethod
    def _storage_summary(medium: StorageMedium) -> str:
        capacity = ""
        if medium.capacity_value:
            capacity = (
                f"{AssetService._display_value(medium.capacity_value)}"
                f"{medium.capacity_unit}"
            )
        return " / ".join(
            value
            for value in (
                medium.medium_type,
                medium.name,
                medium.brand,
                capacity,
                medium.model,
                medium.serial_number,
            )
            if value
        )

    @staticmethod
    def _new_id(prefix: str) -> str:
        return f"{prefix}-{uuid4().hex.upper()}"

    @staticmethod
    def _display_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    @staticmethod
    def _comparable(value: Any) -> str:
        return AssetService._display_value(value).strip()

    @staticmethod
    def _format_asset_row(sheet, row_number: int) -> None:
        header_columns = {
            cell.value: cell.column for cell in sheet[1]
        }
        sheet.cell(row=row_number, column=header_columns["金额"]).number_format = '¥#,##0.00'
        for header in ("取得日期", "启用日期", "出厂日期"):
            column = header_columns[header]
            sheet.cell(row=row_number, column=column).number_format = "yyyy-mm-dd"
        for header in ("创建时间", "更新时间"):
            column = header_columns[header]
            sheet.cell(row=row_number, column=column).number_format = "yyyy-mm-dd hh:mm:ss"

    @staticmethod
    def _format_change_row(sheet, row_number: int) -> None:
        sheet.cell(row=row_number, column=8).number_format = "yyyy-mm-dd hh:mm:ss"

    @staticmethod
    def _format_storage_row(sheet, row_number: int) -> None:
        sheet.cell(row=row_number, column=6).number_format = "0.##"
        for column in (11, 12):
            sheet.cell(row=row_number, column=column).number_format = "yyyy-mm-dd hh:mm:ss"
