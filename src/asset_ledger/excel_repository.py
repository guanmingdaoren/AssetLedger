from __future__ import annotations

from datetime import datetime
from pathlib import Path
import shutil
import tempfile
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ASSET_SHEET = "资产台账"
STORAGE_SHEET = "存储介质"
CHANGE_SHEET = "变更历史"
CATEGORY_SHEET = "分类字典"
LOCATION_SHEET = "位置字典"
ENUM_SHEET = "枚举字典"
CONFIG_SHEET = "系统配置"

REQUIRED_SHEETS = [
    ASSET_SHEET,
    STORAGE_SHEET,
    CHANGE_SHEET,
    CATEGORY_SHEET,
    LOCATION_SHEET,
    ENUM_SHEET,
    CONFIG_SHEET,
]

ASSET_HEADERS = [
    "资产唯一标识符",
    "bm编码",
    "设备器材",
    "一级类别",
    "二级类别",
    "产品规格",
    "产品型号",
    "设备序列号",
    "生产厂家",
    "供应商",
    "品牌",
    "取得方式",
    "取得日期",
    "启用日期",
    "计价方式",
    "金额",
    "管理部门",
    "管理人",
    "使用部门",
    "使用人",
    "存放地点",
    "出厂日期",
    "等级",
    "使用状态",
    "是否已打印标签",
    "备注",
    "创建时间",
    "更新时间",
]

STORAGE_HEADERS = [
    "存储介质ID",
    "资产唯一标识符",
    "介质类型",
    "名称/编号",
    "品牌",
    "容量数值",
    "容量单位",
    "型号",
    "序列号",
    "备注",
    "创建时间",
    "更新时间",
]

CHANGE_HEADERS = [
    "变更ID",
    "事件组ID",
    "资产ID",
    "事件类型",
    "变化字段",
    "修改前值",
    "修改后值",
    "修改时间",
    "操作者",
    "修改说明",
]

CATEGORY_HEADERS = ["类别ID", "父类别ID", "类别名称", "排序", "是否启用"]
LOCATION_HEADERS = ["位置ID", "父位置ID", "位置名称", "排序", "是否启用"]
ENUM_HEADERS = ["枚举类型", "枚举值", "排序", "是否启用"]
CONFIG_HEADERS = ["配置项", "配置值", "说明"]

DEFAULT_CATEGORIES = [
    ("CAT-IT", "", "IT 与通信设备", 10, True),
    ("CAT-IT-PC", "CAT-IT", "电脑", 11, True),
    ("CAT-IT-MONITOR", "CAT-IT", "显示器", 12, True),
    ("CAT-IT-PRINTER", "CAT-IT", "打印机", 13, True),
    ("CAT-IT-NETWORK", "CAT-IT", "网络设备", 14, True),
    ("CAT-OFFICE", "", "办公设备", 20, True),
    ("CAT-OFFICE-PROJECTOR", "CAT-OFFICE", "投影仪", 21, True),
    ("CAT-OFFICE-SCANNER", "CAT-OFFICE", "扫描仪", 22, True),
    ("CAT-OFFICE-COPIER", "CAT-OFFICE", "复印机", 23, True),
    ("CAT-INSTRUMENT", "", "仪器与检测设备", 30, True),
    ("CAT-INSTRUMENT-MEASURE", "CAT-INSTRUMENT", "测量仪器", 31, True),
    ("CAT-INSTRUMENT-LAB", "CAT-INSTRUMENT", "实验仪器", 32, True),
    ("CAT-TOOLS", "", "生产与作业工具", 40, True),
    ("CAT-TOOLS-POWER", "CAT-TOOLS", "电动工具", 41, True),
    ("CAT-TOOLS-HAND", "CAT-TOOLS", "手动工具", 42, True),
    ("CAT-SAFETY", "", "安防与消防设备", 50, True),
    ("CAT-FURNITURE", "", "家具与辅助设施", 60, True),
    ("CAT-OTHER", "", "其他设备", 70, True),
]

DEFAULT_LOCATIONS = [
    ("LOC-WAREHOUSE", "", "仓库", 10, True),
    ("LOC-OFFICE", "", "办公室", 20, True),
    ("LOC-WORKSHOP", "", "作业区", 30, True),
]

DEFAULT_ENUMS = [
    *[("使用状态", value, index * 10, True) for index, value in enumerate(
        ["入库", "在用", "闲置", "维修", "封存", "报废", "丢失"], start=1
    )],
    ("取得方式", "采购", 10, True),
    ("取得方式", "调拨", 20, True),
    ("取得方式", "捐赠", 30, True),
    ("取得方式", "其他", 40, True),
]

DEFAULT_CONFIG = [
    ("工作簿版本", "2", "用于识别工作簿结构"),
    ("下一个资产序号", "1", "生成资产 ID 使用"),
    ("默认操作者", "管理员", "记录变更时使用"),
]

V1_ASSET_HEADERS = [
    "资产ID",
    "装备编码",
    "设备名称",
    "一级类别",
    "二级类别",
    "品牌",
    "型号",
    "序列号",
    "来源",
    "价格",
    "购入日期",
    "启用日期",
    "状态",
    "位置",
    "责任部门",
    "责任人",
    "备注",
    "创建时间",
    "更新时间",
]

V1_TO_V2_HEADERS = {
    "资产ID": "资产唯一标识符",
    "装备编码": "bm编码",
    "设备名称": "设备器材",
    "型号": "产品型号",
    "序列号": "设备序列号",
    "来源": "取得方式",
    "价格": "金额",
    "购入日期": "取得日期",
    "状态": "使用状态",
    "位置": "存放地点",
    "责任部门": "管理部门",
    "责任人": "管理人",
}


class WorkbookValidationError(ValueError):
    pass


class WorkbookLockedError(PermissionError):
    pass


class ExcelRepository:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def initialize(self) -> None:
        if self.path.exists():
            self._migrate_if_needed()
            return
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._create_sheet(workbook, ASSET_SHEET, ASSET_HEADERS, [], "AssetsTable")
        self._create_sheet(workbook, STORAGE_SHEET, STORAGE_HEADERS, [], "StorageMediaTable")
        self._create_sheet(workbook, CHANGE_SHEET, CHANGE_HEADERS, [], "ChangesTable")
        self._create_sheet(
            workbook, CATEGORY_SHEET, CATEGORY_HEADERS, DEFAULT_CATEGORIES, "CategoriesTable"
        )
        self._create_sheet(
            workbook, LOCATION_SHEET, LOCATION_HEADERS, DEFAULT_LOCATIONS, "LocationsTable"
        )
        self._create_sheet(workbook, ENUM_SHEET, ENUM_HEADERS, DEFAULT_ENUMS, "EnumsTable")
        self._create_sheet(
            workbook, CONFIG_SHEET, CONFIG_HEADERS, DEFAULT_CONFIG, "ConfigTable"
        )
        self._save_atomic(workbook, create_backup=False)

    def validate_workbook(self) -> list[str]:
        if not self.path.exists():
            return ["工作簿不存在"]
        try:
            workbook = load_workbook(self.path, read_only=True, data_only=False)
        except Exception as exc:
            return [f"无法读取工作簿：{exc}"]

        errors: list[str] = []
        expected_headers = {
            ASSET_SHEET: ASSET_HEADERS,
            STORAGE_SHEET: STORAGE_HEADERS,
            CHANGE_SHEET: CHANGE_HEADERS,
            CATEGORY_SHEET: CATEGORY_HEADERS,
            LOCATION_SHEET: LOCATION_HEADERS,
            ENUM_SHEET: ENUM_HEADERS,
            CONFIG_SHEET: CONFIG_HEADERS,
        }
        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in workbook.sheetnames:
                errors.append(f"缺少工作表：{sheet_name}")
                continue
            actual = [cell.value for cell in next(workbook[sheet_name].iter_rows(max_row=1))]
            if actual != expected_headers[sheet_name]:
                errors.append(f"工作表字段不正确：{sheet_name}")

        if ASSET_SHEET in workbook.sheetnames:
            ids: set[str] = set()
            codes: set[str] = set()
            for row_number, row in enumerate(
                workbook[ASSET_SHEET].iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(value is not None for value in row):
                    continue
                asset_id = str(row[0] or "")
                equipment_code = str(row[1] or "")
                name = str(row[2] or "")
                if not asset_id:
                    errors.append(f"资产唯一标识符为空：第{row_number}行")
                elif asset_id in ids:
                    errors.append(f"资产唯一标识符重复：{asset_id}")
                if not name:
                    errors.append(f"设备器材为空：第{row_number}行")
                if equipment_code and equipment_code in codes:
                    errors.append(f"bm编码重复：{equipment_code}")
                if asset_id:
                    ids.add(asset_id)
                if equipment_code:
                    codes.add(equipment_code)
        if STORAGE_SHEET in workbook.sheetnames:
            storage_ids: set[str] = set()
            for row_number, row in enumerate(
                workbook[STORAGE_SHEET].iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(value is not None for value in row):
                    continue
                storage_id = str(row[0] or "")
                asset_id = str(row[1] or "")
                if not storage_id:
                    errors.append(f"存储介质ID为空：第{row_number}行")
                elif storage_id in storage_ids:
                    errors.append(f"存储介质ID重复：{storage_id}")
                if not asset_id or asset_id not in ids:
                    errors.append(f"存储介质关联资产不存在：第{row_number}行")
                if storage_id:
                    storage_ids.add(storage_id)
        workbook.close()
        return errors

    def _migrate_if_needed(self) -> None:
        try:
            workbook = load_workbook(self.path)
        except Exception as exc:
            raise WorkbookValidationError(f"无法读取工作簿：{exc}") from exc
        try:
            if CONFIG_SHEET not in workbook.sheetnames:
                return
            version = ""
            for row in workbook[CONFIG_SHEET].iter_rows(min_row=2, values_only=True):
                if row[0] == "工作簿版本":
                    version = str(row[1] or "")
                    break
            if version == "2":
                return
            if version != "1":
                raise WorkbookValidationError(f"不支持的工作簿版本：{version or '未知'}")
            self._migrate_v1(workbook)
        finally:
            workbook.close()

    def _migrate_v1(self, workbook) -> None:
        required_v1 = [
            ASSET_SHEET,
            CHANGE_SHEET,
            CATEGORY_SHEET,
            LOCATION_SHEET,
            ENUM_SHEET,
            CONFIG_SHEET,
        ]
        missing = [name for name in required_v1 if name not in workbook.sheetnames]
        if missing:
            raise WorkbookValidationError(f"旧版工作簿缺少工作表：{', '.join(missing)}")
        expected_v1_headers = {
            ASSET_SHEET: V1_ASSET_HEADERS,
            CHANGE_SHEET: CHANGE_HEADERS,
            CATEGORY_SHEET: CATEGORY_HEADERS,
            LOCATION_SHEET: LOCATION_HEADERS,
            ENUM_SHEET: ENUM_HEADERS,
            CONFIG_SHEET: CONFIG_HEADERS,
        }
        for sheet_name, expected in expected_v1_headers.items():
            actual = [
                cell.value
                for cell in next(workbook[sheet_name].iter_rows(max_row=1))
            ]
            if actual != expected:
                raise WorkbookValidationError(f"旧版工作表字段结构无法识别：{sheet_name}")

        old_rows = [
            dict(zip(V1_ASSET_HEADERS, row))
            for row in workbook[ASSET_SHEET].iter_rows(min_row=2, values_only=True)
            if any(value is not None for value in row)
        ]
        migrated_rows = []
        for old in old_rows:
            migrated = {header: None for header in ASSET_HEADERS}
            for old_header, value in old.items():
                migrated[V1_TO_V2_HEADERS.get(old_header, old_header)] = value
            migrated_rows.append([migrated[header] for header in ASSET_HEADERS])

        del workbook[ASSET_SHEET]
        self._create_sheet(workbook, ASSET_SHEET, ASSET_HEADERS, migrated_rows, "AssetsTable")
        asset_sheet = workbook[ASSET_SHEET]
        workbook._sheets.remove(asset_sheet)
        workbook._sheets.insert(0, asset_sheet)

        self._create_sheet(workbook, STORAGE_SHEET, STORAGE_HEADERS, [], "StorageMediaTable")
        storage_sheet = workbook[STORAGE_SHEET]
        workbook._sheets.remove(storage_sheet)
        workbook._sheets.insert(1, storage_sheet)

        for row in workbook[CHANGE_SHEET].iter_rows(min_row=2):
            if row[4].value in V1_TO_V2_HEADERS:
                row[4].value = V1_TO_V2_HEADERS[row[4].value]
            if row[3].value == "状态变更":
                row[3].value = "使用状态变更"
            elif row[3].value == "位置变更":
                row[3].value = "存放地点变更"
        for row in workbook[ENUM_SHEET].iter_rows(min_row=2):
            if row[0].value == "状态":
                row[0].value = "使用状态"
            elif row[0].value == "来源":
                row[0].value = "取得方式"
        for row in workbook[CONFIG_SHEET].iter_rows(min_row=2):
            if row[0].value == "工作簿版本":
                row[1].value = "2"
                break

        self.backup_workbook()
        self._save_atomic(workbook, create_backup=False)

    def require_valid(self) -> None:
        errors = self.validate_workbook()
        if errors:
            raise WorkbookValidationError("\n".join(errors))

    def load(self, *, data_only: bool = False):
        self.require_valid()
        return load_workbook(self.path, data_only=data_only)

    def backup_workbook(self) -> Path | None:
        if not self.path.exists():
            return None
        backup_dir = self.path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        backup_path = backup_dir / f"{self.path.stem}-{timestamp}{self.path.suffix}"
        shutil.copy2(self.path, backup_path)
        return backup_path

    def save(self, workbook, *, create_backup: bool = True) -> None:
        self._assert_not_locked()
        if create_backup:
            self.backup_workbook()
        self._save_atomic(workbook, create_backup=False)

    def _save_atomic(self, workbook, *, create_backup: bool = False) -> None:
        if create_backup:
            self.backup_workbook()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                dir=self.path.parent, suffix=".xlsx", delete=False
            ) as temporary_file:
                temporary_path = Path(temporary_file.name)
            workbook.save(temporary_path)
            temporary_path.replace(self.path)
        except PermissionError as exc:
            raise WorkbookLockedError("Excel 文件正在被占用，无法保存。") from exc
        finally:
            if temporary_path and temporary_path.exists():
                temporary_path.unlink(missing_ok=True)

    def _assert_not_locked(self) -> None:
        if not self.path.exists():
            return
        try:
            with self.path.open("r+b"):
                pass
        except PermissionError as exc:
            raise WorkbookLockedError("Excel 文件正在被占用，请关闭 Excel 后重试。") from exc

    @staticmethod
    def _create_sheet(
        workbook: Workbook,
        name: str,
        headers: list[str],
        rows: Iterable[Iterable[Any]],
        table_name: str,
    ) -> None:
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        row_count = 1
        for row in rows:
            sheet.append(list(row))
            row_count += 1

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{ExcelRepository._column_letter(len(headers))}{row_count}"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = {
            "资产唯一标识符": 18,
            "bm编码": 18,
            "设备器材": 20,
            "名称/编号": 18,
            "备注": 30,
            "修改说明": 30,
            "创建时间": 20,
            "更新时间": 20,
            "修改时间": 20,
        }
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[ExcelRepository._column_letter(index)].width = widths.get(
                header, 16
            )

        # Excel tables require at least one data row.
        if row_count == 1:
            sheet.append([None] * len(headers))
            row_count = 2
        table = Table(
            displayName=table_name,
            ref=f"A1:{ExcelRepository._column_letter(len(headers))}{row_count}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    @staticmethod
    def _column_letter(index: int) -> str:
        result = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
