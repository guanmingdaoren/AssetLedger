from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from asset_ledger.excel_repository import (
    ASSET_HEADERS,
    CATEGORY_HEADERS,
    CHANGE_HEADERS,
    CONFIG_HEADERS,
    DEFAULT_CATEGORIES,
    DEFAULT_CONFIG,
    DEFAULT_ENUMS,
    DEFAULT_LOCATIONS,
    ENUM_HEADERS,
    ExcelRepository,
    LOCATION_HEADERS,
    STORAGE_HEADERS,
    WorkbookChangedExternallyError,
    WorkbookValidationError,
    WorkbookLockedError,
)


V2_ASSET_HEADERS = [
    "资产唯一标识符",
    "bm编码",
    "设备器材",
    "一级类别",
    "二级类别",
    "产品规格",
    "产品型号",
    "设备序列号",
    "生产厂家",
    "供应商",
    "品牌",
    "取得方式",
    "取得日期",
    "启用日期",
    "计价方式",
    "金额",
    "管理部门",
    "管理人",
    "使用部门",
    "使用人",
    "存放地点",
    "出厂日期",
    "等级",
    "使用状态",
    "是否已打印标签",
    "备注",
    "创建时间",
    "更新时间",
]

V2_STORAGE_HEADERS = [
    "存储介质ID",
    "资产唯一标识符",
    "介质类型",
    "名称/编号",
    "品牌",
    "容量数值",
    "容量单位",
    "型号",
    "序列号",
    "备注",
    "创建时间",
    "更新时间",
]

V2_CHANGE_HEADERS = [
    "变更ID",
    "事件组ID",
    "资产ID",
    "事件类型",
    "变化字段",
    "修改前值",
    "修改后值",
    "修改时间",
    "操作者",
    "修改说明",
]

V3_ASSET_HEADERS = [
    "资产UID",
    "bm编码",
    "资产唯一标识符",
    "设备器材",
    "一级类别",
    "二级类别",
    "产品规格",
    "产品型号",
    "设备序列号",
    "生产厂家",
    "供应商",
    "品牌",
    "取得方式",
    "取得日期",
    "启用日期",
    "计价方式",
    "金额",
    "管理部门",
    "管理人",
    "使用部门",
    "使用人",
    "存放地点",
    "出厂日期",
    "等级",
    "使用状态",
    "是否已打印标签",
    "备注",
    "创建时间",
    "更新时间",
]

V3_STORAGE_HEADERS = [
    "存储介质ID",
    "资产UID",
    "介质类型",
    "名称/编号",
    "品牌",
    "容量数值",
    "容量单位",
    "型号",
    "序列号",
    "备注",
    "创建时间",
    "更新时间",
]

V3_CHANGE_HEADERS = [
    "变更ID",
    "事件组ID",
    "资产UID",
    "事件类型",
    "变化字段",
    "修改前值",
    "修改后值",
    "修改时间",
    "操作者",
    "修改说明",
]

V4_ASSET_HEADERS = [
    "资产UID",
    "bm编码",
    "资产唯一标识符",
    "设备器材",
    "一级类别",
    "二级类别",
    "类别路径",
    "产品规格",
    "产品型号",
    "设备序列号",
    "生产厂家",
    "供应商",
    "品牌",
    "取得方式",
    "取得日期",
    "启用日期",
    "计价方式",
    "金额",
    "管理部门",
    "管理人",
    "使用部门",
    "使用人",
    "存放地点",
    "出厂日期",
    "等级",
    "使用状态",
    "是否已打印标签",
    "备注",
    "创建时间",
    "更新时间",
]

V4_STORAGE_HEADERS = [
    "存储介质ID",
    "资产UID",
    "介质类型",
    "使用状态",
    "名称/编号",
    "品牌",
    "容量数值",
    "容量单位",
    "型号",
    "序列号",
    "备注",
    "创建时间",
    "更新时间",
]

V1_ASSET_HEADERS = [
    "资产ID",
    "装备编码",
    "设备名称",
    "一级类别",
    "二级类别",
    "品牌",
    "型号",
    "序列号",
    "来源",
    "价格",
    "购入日期",
    "启用日期",
    "状态",
    "位置",
    "责任部门",
    "责任人",
    "备注",
    "创建时间",
    "更新时间",
]


class ExcelRepositoryTests(unittest.TestCase):
    def test_initializes_workbook_with_required_sheets_and_defaults(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"

            repository = ExcelRepository(path)
            repository.initialize()

            workbook = load_workbook(path, read_only=True, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "资产台账",
                    "存储介质",
                    "变更历史",
                    "分类字典",
                    "位置字典",
                    "枚举字典",
                    "系统配置",
                ],
            )
            self.assertEqual(workbook["系统配置"]["B2"].value, "5")
            self.assertEqual(
                [cell.value for cell in next(workbook["资产台账"].iter_rows(max_row=1))],
                ASSET_HEADERS,
            )
            self.assertIn("备注1", ASSET_HEADERS)
            self.assertIn("备注2", ASSET_HEADERS)
            self.assertNotIn("备注", ASSET_HEADERS)
            self.assertEqual(
                ASSET_HEADERS[:7],
                ["资产UID", "bm编码", "资产唯一标识符", "设备器材", "一级类别", "二级类别", "类别路径"],
            )
            self.assertEqual(
                [cell.value for cell in next(workbook["存储介质"].iter_rows(max_row=1))],
                STORAGE_HEADERS,
            )
            self.assertEqual(STORAGE_HEADERS[1], "资产UID")
            self.assertEqual(STORAGE_HEADERS[3], "使用状态")
            self.assertEqual(CHANGE_HEADERS[2], "资产UID")
            self.assertGreater(workbook["分类字典"].max_row, 7)
            self.assertGreater(workbook["枚举字典"].max_row, 7)
            workbook.close()

    def test_initialize_migrates_v1_workbook_and_creates_backup(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            ExcelRepository._create_sheet(
                workbook,
                "资产台账",
                V1_ASSET_HEADERS,
                [
                    [
                        "AST-2026-000001",
                        "BM-001",
                        "旧电脑",
                        "IT 与通信设备",
                        "电脑",
                        "联想",
                        "T14",
                        "SN-001",
                        "采购",
                        5999,
                        "2026-01-02",
                        "2026-01-05",
                        "在用",
                        "办公室",
                        "技术部",
                        "张三",
                        "旧备注",
                        "2026-01-02 10:00:00",
                        "2026-01-02 10:00:00",
                    ]
                ],
                "AssetsTable",
            )
            ExcelRepository._create_sheet(
                workbook, "变更历史", V2_CHANGE_HEADERS, [], "ChangesTable"
            )
            ExcelRepository._create_sheet(
                workbook,
                "分类字典",
                CATEGORY_HEADERS,
                DEFAULT_CATEGORIES,
                "CategoriesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "位置字典",
                LOCATION_HEADERS,
                DEFAULT_LOCATIONS,
                "LocationsTable",
            )
            ExcelRepository._create_sheet(
                workbook, "枚举字典", ENUM_HEADERS, DEFAULT_ENUMS, "EnumsTable"
            )
            v1_config = [
                ("工作簿版本", "1", "用于识别工作簿结构"),
                *[row for row in DEFAULT_CONFIG if row[0] != "工作簿版本"],
            ]
            ExcelRepository._create_sheet(
                workbook, "系统配置", CONFIG_HEADERS, v1_config, "ConfigTable"
            )
            workbook.save(path)
            workbook.close()

            repository = ExcelRepository(path)
            repository.initialize()

            migrated = load_workbook(path, read_only=True, data_only=True)
            self.assertIn("存储介质", migrated.sheetnames)
            self.assertEqual(migrated["系统配置"]["B2"].value, "5")
            row = next(migrated["资产台账"].iter_rows(min_row=2, values_only=True))
            headers = [
                cell.value for cell in next(migrated["资产台账"].iter_rows(max_row=1))
            ]
            values = dict(zip(headers, row))
            self.assertEqual(values["资产UID"], "AST-2026-000001")
            self.assertIsNone(values["资产唯一标识符"])
            self.assertEqual(values["bm编码"], "BM-001")
            self.assertEqual(values["设备器材"], "旧电脑")
            self.assertEqual(values["类别路径"], "IT 与通信设备 / 电脑")
            self.assertEqual(values["产品型号"], "T14")
            self.assertEqual(values["管理部门"], "技术部")
            self.assertEqual(values["管理人"], "张三")
            self.assertEqual(values["备注1"], "旧备注")
            self.assertIsNone(values["备注2"])
            migrated.close()
            self.assertEqual(len(list((path.parent / "backups").glob("assets-*.xlsx"))), 1)
            self.assertEqual(repository.validate_workbook(), [])

    def test_initialize_migrates_v2_workbook_to_asset_uid_structure(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            ExcelRepository._create_sheet(
                workbook,
                "资产台账",
                V2_ASSET_HEADERS,
                [
                    [
                        "AST-2026-000001",
                        "BM-001",
                        "旧电脑",
                        "IT 与通信设备",
                        "电脑",
                        "14英寸 / 32GB",
                        "T14",
                        "SN-001",
                        "联想（北京）有限公司",
                        "示例供应商",
                        "联想",
                        "采购",
                        "2026-01-02",
                        "2026-01-05",
                        "原值",
                        5999,
                        "技术部",
                        "张三",
                        "研发部",
                        "李四",
                        "办公室",
                        "2025-12-01",
                        "A",
                        "在用",
                        True,
                        "旧备注",
                        "2026-01-02 10:00:00",
                        "2026-01-02 10:00:00",
                    ]
                ],
                "AssetsTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "存储介质",
                V2_STORAGE_HEADERS,
                [
                    [
                        "MED-1",
                        "AST-2026-000001",
                        "SSD",
                        "系统盘",
                        "三星",
                        1,
                        "TB",
                        "990 PRO",
                        "SSD-SN-001",
                        "",
                        "2026-01-02 10:00:00",
                        "2026-01-02 10:00:00",
                    ]
                ],
                "StorageMediaTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "变更历史",
                V2_CHANGE_HEADERS,
                [
                    [
                        "CHG-1",
                        "EVT-1",
                        "AST-2026-000001",
                        "信息修改",
                        "资产唯一标识符",
                        "",
                        "AST-2026-000001",
                        "2026-01-02 10:00:00",
                        "管理员",
                        "",
                    ]
                ],
                "ChangesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "分类字典",
                CATEGORY_HEADERS,
                DEFAULT_CATEGORIES,
                "CategoriesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "位置字典",
                LOCATION_HEADERS,
                DEFAULT_LOCATIONS,
                "LocationsTable",
            )
            ExcelRepository._create_sheet(
                workbook, "枚举字典", ENUM_HEADERS, DEFAULT_ENUMS, "EnumsTable"
            )
            v2_config = [
                ("工作簿版本", "2", "用于识别工作簿结构"),
                *[row for row in DEFAULT_CONFIG if row[0] != "工作簿版本"],
            ]
            ExcelRepository._create_sheet(
                workbook, "系统配置", CONFIG_HEADERS, v2_config, "ConfigTable"
            )
            workbook.save(path)
            workbook.close()

            repository = ExcelRepository(path)
            repository.initialize()

            migrated = load_workbook(path, read_only=True, data_only=True)
            asset_headers = [
                cell.value for cell in next(migrated["资产台账"].iter_rows(max_row=1))
            ]
            asset_values = dict(
                zip(asset_headers, next(migrated["资产台账"].iter_rows(min_row=2, values_only=True)))
            )
            storage_headers = [
                cell.value for cell in next(migrated["存储介质"].iter_rows(max_row=1))
            ]
            change_headers = [
                cell.value for cell in next(migrated["变更历史"].iter_rows(max_row=1))
            ]
            change_values = dict(
                zip(change_headers, next(migrated["变更历史"].iter_rows(min_row=2, values_only=True)))
            )

            self.assertEqual(migrated["系统配置"]["B2"].value, "5")
            self.assertEqual(asset_headers, ASSET_HEADERS)
            self.assertEqual(asset_values["资产UID"], "AST-2026-000001")
            self.assertIsNone(asset_values["资产唯一标识符"])
            self.assertEqual(asset_values["类别路径"], "IT 与通信设备 / 电脑")
            self.assertEqual(asset_values["bm编码"], "BM-001")
            self.assertEqual(asset_values["备注1"], "旧备注")
            self.assertIsNone(asset_values["备注2"])
            self.assertEqual(storage_headers[1], "资产UID")
            self.assertEqual(storage_headers[3], "使用状态")
            self.assertEqual(change_headers[2], "资产UID")
            self.assertEqual(change_values["资产UID"], "AST-2026-000001")
            self.assertEqual(change_values["变化字段"], "资产UID")
            migrated.close()
            self.assertEqual(len(list((path.parent / "backups").glob("assets-*.xlsx"))), 1)
            self.assertEqual(repository.validate_workbook(), [])

    def test_initialize_migrates_v3_workbook_to_category_path_and_storage_status(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            ExcelRepository._create_sheet(
                workbook,
                "资产台账",
                V3_ASSET_HEADERS,
                [
                    [
                        "AST-2026-000001",
                        "BM-001",
                        "ZC-001",
                        "旧电脑",
                        "IT 与通信设备",
                        "电脑",
                        "14英寸 / 32GB",
                        "T14",
                        "SN-001",
                        "联想（北京）有限公司",
                        "示例供应商",
                        "联想",
                        "采购",
                        "2026-01-02",
                        "2026-01-05",
                        "原值",
                        5999,
                        "技术部",
                        "张三",
                        "研发部",
                        "李四",
                        "办公室",
                        "2025-12-01",
                        "A",
                        "在用",
                        True,
                        "旧备注",
                        "2026-01-02 10:00:00",
                        "2026-01-02 10:00:00",
                    ]
                ],
                "AssetsTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "存储介质",
                V3_STORAGE_HEADERS,
                [
                    [
                        "MED-1",
                        "AST-2026-000001",
                        "SSD",
                        "系统盘",
                        "三星",
                        1,
                        "TB",
                        "990 PRO",
                        "SSD-SN-001",
                        "",
                        "2026-01-02 10:00:00",
                        "2026-01-02 10:00:00",
                    ]
                ],
                "StorageMediaTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "变更历史",
                V3_CHANGE_HEADERS,
                [],
                "ChangesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "分类字典",
                CATEGORY_HEADERS,
                DEFAULT_CATEGORIES,
                "CategoriesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "位置字典",
                LOCATION_HEADERS,
                DEFAULT_LOCATIONS,
                "LocationsTable",
            )
            ExcelRepository._create_sheet(
                workbook, "枚举字典", ENUM_HEADERS, DEFAULT_ENUMS, "EnumsTable"
            )
            v3_config = [
                ("工作簿版本", "3", "用于识别工作簿结构"),
                *[row for row in DEFAULT_CONFIG if row[0] != "工作簿版本"],
            ]
            ExcelRepository._create_sheet(
                workbook, "系统配置", CONFIG_HEADERS, v3_config, "ConfigTable"
            )
            workbook.save(path)
            workbook.close()

            repository = ExcelRepository(path)
            repository.initialize()

            migrated = load_workbook(path, read_only=True, data_only=True)
            asset_headers = [
                cell.value for cell in next(migrated["资产台账"].iter_rows(max_row=1))
            ]
            storage_headers = [
                cell.value for cell in next(migrated["存储介质"].iter_rows(max_row=1))
            ]
            asset_values = dict(
                zip(asset_headers, next(migrated["资产台账"].iter_rows(min_row=2, values_only=True)))
            )
            storage_values = dict(
                zip(storage_headers, next(migrated["存储介质"].iter_rows(min_row=2, values_only=True)))
            )

            self.assertEqual(migrated["系统配置"]["B2"].value, "5")
            self.assertEqual(asset_headers, ASSET_HEADERS)
            self.assertEqual(storage_headers, STORAGE_HEADERS)
            self.assertEqual(asset_values["类别路径"], "IT 与通信设备 / 电脑")
            self.assertEqual(asset_values["备注1"], "旧备注")
            self.assertIsNone(asset_values["备注2"])
            self.assertIsNone(storage_values["使用状态"])
            migrated.close()
            self.assertEqual(len(list((path.parent / "backups").glob("assets-*.xlsx"))), 1)
            self.assertEqual(repository.validate_workbook(), [])

    def test_initialize_migrates_v4_workbook_to_two_note_fields(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            ExcelRepository._create_sheet(
                workbook,
                "资产台账",
                V4_ASSET_HEADERS,
                [
                    [
                        "AST-2026-000001",
                        "BM-001",
                        "ZC-001",
                        "旧电脑",
                        "IT 与通信设备",
                        "电脑",
                        "IT 与通信设备 / 电脑",
                        "14英寸 / 32GB",
                        "T14",
                        "SN-001",
                        "联想（北京）有限公司",
                        "示例供应商",
                        "联想",
                        "采购",
                        "2026-01-02",
                        "2026-01-05",
                        "原值",
                        5999,
                        "技术部",
                        "张三",
                        "研发部",
                        "李四",
                        "办公室",
                        "2025-12-01",
                        "A",
                        "在用",
                        True,
                        "旧备注",
                        "2026-01-02 10:00:00",
                        "2026-01-02 10:00:00",
                    ]
                ],
                "AssetsTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "存储介质",
                V4_STORAGE_HEADERS,
                [],
                "StorageMediaTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "变更历史",
                V3_CHANGE_HEADERS,
                [],
                "ChangesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "分类字典",
                CATEGORY_HEADERS,
                DEFAULT_CATEGORIES,
                "CategoriesTable",
            )
            ExcelRepository._create_sheet(
                workbook,
                "位置字典",
                LOCATION_HEADERS,
                DEFAULT_LOCATIONS,
                "LocationsTable",
            )
            ExcelRepository._create_sheet(
                workbook, "枚举字典", ENUM_HEADERS, DEFAULT_ENUMS, "EnumsTable"
            )
            v4_config = [
                ("工作簿版本", "4", "用于识别工作簿结构"),
                *[row for row in DEFAULT_CONFIG if row[0] != "工作簿版本"],
            ]
            ExcelRepository._create_sheet(
                workbook, "系统配置", CONFIG_HEADERS, v4_config, "ConfigTable"
            )
            workbook.save(path)
            workbook.close()

            repository = ExcelRepository(path)
            repository.initialize()

            migrated = load_workbook(path, read_only=True, data_only=True)
            asset_headers = [
                cell.value for cell in next(migrated["资产台账"].iter_rows(max_row=1))
            ]
            asset_values = dict(
                zip(asset_headers, next(migrated["资产台账"].iter_rows(min_row=2, values_only=True)))
            )

            self.assertEqual(migrated["系统配置"]["B2"].value, "5")
            self.assertEqual(asset_headers, ASSET_HEADERS)
            self.assertEqual(asset_values["备注1"], "旧备注")
            self.assertIsNone(asset_values["备注2"])
            migrated.close()
            self.assertEqual(len(list((path.parent / "backups").glob("assets-*.xlsx"))), 1)
            self.assertEqual(repository.validate_workbook(), [])

    def test_validate_workbook_reports_missing_sheet(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()

            workbook = load_workbook(path)
            del workbook["变更历史"]
            workbook.save(path)

            errors = repository.validate_workbook()

            self.assertIn("缺少工作表：变更历史", errors)

    def test_validate_workbook_reports_missing_asset_sheet_without_crashing(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()

            workbook = load_workbook(path)
            workbook["存储介质"].append(
                ["MED-1", "AST-2026-000001", "SSD", "", "", 512, "GB"]
            )
            del workbook["资产台账"]
            workbook.save(path)
            workbook.close()

            errors = repository.validate_workbook()

            self.assertIn("缺少工作表：资产台账", errors)

    def test_validate_workbook_reports_wrong_headers_and_duplicate_ids(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()
            workbook = load_workbook(path)
            sheet = workbook["资产台账"]
            sheet["C1"] = "错误字段"
            sheet.append(["AST-2026-000001", "EQ-1"])
            sheet.append(["AST-2026-000001", "EQ-2"])
            sheet.append(["", "", "无编码设备"])
            workbook.save(path)
            workbook.close()

            errors = repository.validate_workbook()

            self.assertIn("工作表字段不正确：资产台账", errors)
            self.assertIn("资产UID重复：AST-2026-000001", errors)
            self.assertIn("资产UID为空：第4行", errors)
            self.assertFalse(any("bm编码为空" in error for error in errors))

    def test_failed_atomic_save_preserves_original_workbook(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()
            workbook = load_workbook(path)

            with patch.object(workbook, "save", side_effect=RuntimeError("写入失败")):
                with self.assertRaisesRegex(RuntimeError, "写入失败"):
                    repository.save(workbook, create_backup=False)
            workbook.close()

            self.assertEqual(repository.validate_workbook(), [])

    def test_locked_workbook_prevents_save(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()
            workbook = load_workbook(path)

            with patch.object(Path, "open", side_effect=PermissionError("locked")):
                with self.assertRaises(WorkbookLockedError):
                    repository.save(workbook, create_backup=False)
            workbook.close()

    def test_fingerprint_detects_external_workbook_change_before_save(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()
            expected = repository.fingerprint()
            workbook = repository.load()

            external = load_workbook(path)
            external["系统配置"]["B4"] = "外部修改"
            external.save(path)
            external.close()

            with self.assertRaises(WorkbookChangedExternallyError):
                repository.save(
                    workbook,
                    create_backup=False,
                    expected_fingerprint=expected,
                )
            workbook.close()

            saved = load_workbook(path, read_only=True, data_only=True)
            self.assertEqual(saved["系统配置"]["B4"].value, "外部修改")
            saved.close()

    def test_external_change_during_temporary_save_is_not_overwritten(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()
            expected = repository.fingerprint()
            workbook = repository.load()
            original_save = workbook.save

            def save_then_modify_external_file(temporary_path) -> None:
                original_save(temporary_path)
                external = load_workbook(path)
                external["系统配置"]["B4"] = "保存期间外部修改"
                external.save(path)
                external.close()

            with patch.object(workbook, "save", side_effect=save_then_modify_external_file):
                with self.assertRaises(WorkbookChangedExternallyError):
                    repository.save(
                        workbook,
                        create_backup=False,
                        expected_fingerprint=expected,
                    )
            workbook.close()

            saved = load_workbook(path, read_only=True, data_only=True)
            self.assertEqual(saved["系统配置"]["B4"].value, "保存期间外部修改")
            saved.close()

    def test_failed_consistent_load_closes_open_workbook(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "assets.xlsx"
            repository = ExcelRepository(path)
            repository.initialize()
            expected = repository.fingerprint()
            workbook = load_workbook(path)

            with patch(
                "asset_ledger.excel_repository.load_workbook", return_value=workbook
            ), patch.object(
                repository,
                "fingerprint",
                side_effect=[expected, WorkbookValidationError("读取后文件消失")],
            ), patch.object(workbook, "close", wraps=workbook.close) as close:
                with self.assertRaisesRegex(WorkbookValidationError, "读取后文件消失"):
                    repository.load_with_fingerprint()

            close.assert_called_once()


if __name__ == "__main__":
    unittest.main()
